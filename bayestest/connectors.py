from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "variant": ("variant", "variant_name", "group", "arm", "treatment", "bucket", "cohort", "experience"),
    "visitors": ("visitors", "users", "sessions", "visits", "exposures", "trials", "n"),
    "conversions": ("conversions", "orders", "purchases", "signups", "successes", "click_sessions", "clicks"),
    "is_control": ("is_control", "control", "is_baseline"),
    "revenue_sum": ("revenue_sum", "sum_revenue", "gross_revenue_sum"),
    "revenue_sum_squares": ("revenue_sum_squares", "sum_revenue_squares", "revenue_ss", "revenue_sq_sum"),
}


def read_table(path: str, sheet: str | None = None) -> list[dict[str, Any]]:
    p = Path(path)
    suffix = p.suffix.lower()

    if suffix == ".csv":
        with p.open("r", encoding="utf-8", newline="") as f:
            return list(csv.DictReader(f))

    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        out: list[dict[str, Any]] = []
        for row in rows[1:]:
            obj: dict[str, Any] = {}
            for i, key in enumerate(headers):
                if not key:
                    continue
                obj[key] = row[i] if i < len(row) else None
            out.append(obj)
        return out

    raise ValueError("Unsupported file type. Use .csv or .xlsx/.xlsm")


def _to_int(value: Any, field_name: str) -> int:
    if value is None or value == "":
        raise ValueError(f"Missing required integer field '{field_name}'.")
    return int(float(value))


def _to_float_or_none(value: Any) -> float | None:
    if value is None or value == "":
        return None
    return float(value)


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "control"}


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def infer_columns(rows: list[dict[str, Any]], primary_metric: str = "conversion_rate") -> dict[str, str]:
    if not rows:
        raise ValueError("Input table is empty.")

    headers = {str(key): _normalize_header(key) for key in rows[0].keys()}
    inferred: dict[str, str] = {}

    for field in ("variant", "visitors", "conversions", "is_control", "revenue_sum", "revenue_sum_squares"):
        if field in {"revenue_sum", "revenue_sum_squares"} and primary_metric != "arpu":
            continue
        for original, normalized in headers.items():
            if normalized in COLUMN_ALIASES[field]:
                inferred[field] = original
                break

    return inferred


def detect_primary_metric(rows: list[dict[str, Any]], mapping: dict[str, Any] | None = None) -> str:
    if mapping and mapping.get("primary_metric"):
        return str(mapping["primary_metric"])

    inferred = infer_columns(rows, primary_metric="arpu")
    if "revenue_sum" in inferred and "revenue_sum_squares" in inferred:
        return "arpu"
    return "conversion_rate"


def merge_analysis_config(
    *,
    mapping: dict[str, Any] | None,
    defaults: dict[str, Any] | None,
) -> dict[str, Any]:
    config = dict(mapping or {})
    for key, value in (defaults or {}).items():
        if value is not None:
            config[key] = value
    return config


def _require_column(rows: list[dict[str, Any]], column_name: str, semantic_name: str) -> None:
    if rows and column_name not in rows[0]:
        aliases = ", ".join(COLUMN_ALIASES.get(semantic_name, ()))
        raise ValueError(
            f"Could not find a column for '{semantic_name}'. "
            f"Pass a mapping file or rename the source column. Tried aliases: {aliases}."
        )


def build_payload_from_rows(
    rows: list[dict[str, Any]],
    mapping: dict[str, Any] | None = None,
    defaults: dict[str, Any] | None = None,
) -> dict[str, Any]:
    config = merge_analysis_config(mapping=mapping, defaults=defaults)
    primary_metric = str(config.get("primary_metric") or detect_primary_metric(rows, mapping)).strip().lower()
    inferred_columns = infer_columns(rows, primary_metric=primary_metric)
    columns = config.get("columns", {})
    name_col = columns.get("variant") or inferred_columns.get("variant") or "variant"
    visitors_col = columns.get("visitors") or inferred_columns.get("visitors") or "visitors"
    conversions_col = columns.get("conversions") or inferred_columns.get("conversions") or "conversions"
    is_control_col = columns.get("is_control")
    if not is_control_col:
        is_control_col = inferred_columns.get("is_control")
    revenue_sum_col = columns.get("revenue_sum") or inferred_columns.get("revenue_sum")
    revenue_sum_sq_col = columns.get("revenue_sum_squares") or inferred_columns.get("revenue_sum_squares")

    control_rule = config.get("control", {})
    control_col = control_rule.get("column")
    control_value = control_rule.get("value")

    if not control_col and not is_control_col:
        control_col = name_col
        control_value = "control"

    _require_column(rows, name_col, "variant")
    _require_column(rows, visitors_col, "visitors")
    _require_column(rows, conversions_col, "conversions")
    if primary_metric == "arpu":
        if not revenue_sum_col or not revenue_sum_sq_col:
            raise ValueError(
                "ARPU analysis requires revenue_sum and revenue_sum_squares columns. "
                "Pass a mapping file or use matching source headers."
            )
        _require_column(rows, revenue_sum_col, "revenue_sum")
        _require_column(rows, revenue_sum_sq_col, "revenue_sum_squares")

    variants: list[dict[str, Any]] = []
    for row in rows:
        variant_name = str(row.get(name_col, "")).strip()
        if not variant_name:
            continue

        is_control = False
        if is_control_col:
            is_control = _to_bool(row.get(is_control_col))
        elif control_col is not None and control_value is not None:
            is_control = str(row.get(control_col, "")).strip().lower() == str(control_value).strip().lower()

        variant_obj = {
            "name": variant_name,
            "visitors": _to_int(row.get(visitors_col), visitors_col),
            "conversions": _to_int(row.get(conversions_col), conversions_col),
            "is_control": is_control,
        }

        if revenue_sum_col:
            variant_obj["revenue_sum"] = _to_float_or_none(row.get(revenue_sum_col))
        if revenue_sum_sq_col:
            variant_obj["revenue_sum_squares"] = _to_float_or_none(row.get(revenue_sum_sq_col))

        variants.append(variant_obj)

    payload = {
        "experiment_name": config.get("experiment_name", "table_input_experiment"),
        "method": config.get("method", "bayesian"),
        "primary_metric": primary_metric,
        "alpha": config.get("alpha", 0.05),
        "look_index": config.get("look_index", 1),
        "max_looks": config.get("max_looks", 1),
        "information_fraction": config.get("information_fraction"),
        "decision_thresholds": config.get("decision_thresholds"),
        "decision_policy": config.get("decision_policy"),
        "guardrails": config.get("guardrails", []),
        "samples": config.get("samples", 50000),
        "random_seed": config.get("random_seed", 7),
        "variants": variants,
    }

    return payload


def build_duration_request_from_rows(
    rows: list[dict[str, Any]],
    mapping: dict[str, Any],
) -> dict[str, Any]:
    if not rows:
        raise ValueError("Duration input table is empty.")

    row = rows[0]
    columns = mapping.get("columns", {})

    def get_value(key: str, default: Any = None) -> Any:
        col = columns.get(key)
        if col is None:
            return mapping.get(key, default)
        value = row.get(col)
        return mapping.get(key, default) if value in (None, "") else value

    method = str(get_value("method", "frequentist")).strip().lower()
    baseline_rate = float(get_value("baseline_rate"))
    relative_mde = float(get_value("relative_mde"))
    daily_traffic = int(float(get_value("daily_traffic")))

    out = {
        "method": method,
        "baseline_rate": baseline_rate,
        "relative_mde": relative_mde,
        "daily_traffic": daily_traffic,
        "n_variants": int(float(get_value("n_variants", 2))),
        "alpha": float(get_value("alpha", 0.05)),
        "power": float(get_value("power", 0.8)),
        "max_looks": int(float(get_value("max_looks", 10))),
        "prob_threshold": float(get_value("prob_threshold", 0.95)),
        "max_expected_loss": float(get_value("max_expected_loss", 0.001)),
        "assurance_target": float(get_value("assurance_target", 0.8)),
        "max_days": int(float(get_value("max_days", 60))),
    }
    return out
