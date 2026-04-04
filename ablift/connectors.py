from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_table(path: str, sheet: str | None = None) -> list[dict[str, Any]]:
    p = Path(path)
    suffix = p.suffix.lower()

    if suffix == ".csv":
        with p.open("r", encoding="utf-8", newline="") as f:
            return list(csv.DictReader(f))

    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        out: list[dict[str, Any]] = []
        for row in rows[1:]:
            obj: dict[str, Any] = {}
            for i, key in enumerate(headers):
                if not key:
                    continue
                obj[key] = row[i] if i < len(row) else None
            out.append(obj)
        return out

    raise ValueError("Unsupported file type. Use .csv or .xlsx/.xlsm")


def _to_int(value: Any, field_name: str) -> int:
    if value is None or value == "":
        raise ValueError(f"Missing required integer field '{field_name}'.")
    return int(float(value))


def _to_float_or_none(value: Any) -> float | None:
    if value is None or value == "":
        return None
    return float(value)


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "control"}


_BOOL_STRINGS = {"0", "1", "true", "false", "yes", "no", "y", "n"}


def _is_binary_col(values: list[Any]) -> bool:
    """Return True if all non-None values are 0, 1, boolean, or boolean-string."""
    for v in values:
        if v is None or v == "":
            continue
        if isinstance(v, bool):
            continue
        if str(v).strip().lower() in _BOOL_STRINGS:
            continue
        try:
            f = float(v)
            if f not in (0.0, 1.0):
                return False
        except (ValueError, TypeError):
            return False
    return True


def _find_control_column(rows: list[dict[str, Any]], skip_cols: int) -> str | None:
    """Return the name of a boolean-valued column beyond the first skip_cols positional columns."""
    if not rows:
        return None
    all_cols = list(rows[0].keys())
    candidate_cols = all_cols[skip_cols:]
    for col in candidate_cols:
        values = [row.get(col) for row in rows]
        if not _is_binary_col(values):
            continue
        has_true = any(_to_bool(v) for v in values)
        has_false = any(not _to_bool(v) for v in values if v not in (None, ""))
        if has_true and has_false:
            return col
    return None


def detect_input_shape(rows: list[dict[str, Any]]) -> str:
    """Return 'row_level' if col[0] has repeated values, else 'aggregated'."""
    if not rows:
        raise ValueError("Input table is empty.")
    col0 = list(rows[0].keys())[0]
    values = [str(row.get(col0, "")) for row in rows]
    return "row_level" if len(set(values)) < len(values) else "aggregated"


def _build_payload_from_row_level(
    rows: list[dict[str, Any]], defaults: dict[str, Any] | None = None
) -> dict[str, Any]:
    defaults = defaults or {}
    cols = list(rows[0].keys())
    if len(cols) < 2:
        raise ValueError("Row-level input requires at least 2 columns: variant name and outcome.")
    variant_col = cols[0]
    outcome_col = cols[1]
    control_col = _find_control_column(rows, skip_cols=2)

    # Group rows by variant, preserving first-appearance order
    groups: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        name = str(row.get(variant_col, "")).strip()
        if not name:
            continue
        if name not in groups:
            groups[name] = []
        groups[name].append(row)

    if not groups:
        raise ValueError("Row-level input has no variant data.")

    all_outcomes = [row.get(outcome_col) for row in rows if row.get(outcome_col) not in (None, "")]
    is_conversion = _is_binary_col(all_outcomes)
    primary_metric = defaults.get("primary_metric") or (
        "conversion_rate" if is_conversion else "arpu"
    )

    first_variant = next(iter(groups))
    variants: list[dict[str, Any]] = []
    for name, group_rows in groups.items():
        if control_col:
            is_ctrl = any(_to_bool(r.get(control_col)) for r in group_rows)
        else:
            is_ctrl = name == first_variant

        variant_obj: dict[str, Any] = {"name": name, "is_control": is_ctrl}
        if is_conversion:
            variant_obj["visitors"] = len(group_rows)
            variant_obj["conversions"] = sum(
                int(float(r.get(outcome_col, 0))) for r in group_rows
            )
        else:
            values = [float(r.get(outcome_col, 0)) for r in group_rows]
            variant_obj["visitors"] = len(values)
            variant_obj["revenue_sum"] = sum(values)
            variant_obj["revenue_sum_squares"] = sum(v * v for v in values)
        variants.append(variant_obj)

    return _build_final_payload(
        variants, primary_metric, defaults, source_type="row_level", arpu_approximate=False
    )


def _build_payload_from_aggregated(
    rows: list[dict[str, Any]], defaults: dict[str, Any] | None = None
) -> dict[str, Any]:
    defaults = defaults or {}
    cols = list(rows[0].keys())
    if len(cols) < 3:
        raise ValueError(
            "Aggregated input requires at least 3 columns: variant name, visitors, metric."
        )
    variant_col = cols[0]
    visitors_col = cols[1]
    metric_col = cols[2]
    control_col = _find_control_column(rows, skip_cols=3)

    # Determine if col[2] is conversions (integers <= visitors) or revenue_sum
    arpu_approximate = False
    is_conversion = True
    for row in rows:
        try:
            metric_val = float(row.get(metric_col, 0) or 0)
            visitors_val = float(row.get(visitors_col, 0) or 0)
            if metric_val != int(metric_val) or metric_val > visitors_val:
                is_conversion = False
                arpu_approximate = True
                break
        except (ValueError, TypeError):
            pass

    primary_metric = defaults.get("primary_metric") or (
        "conversion_rate" if is_conversion else "arpu"
    )

    variants: list[dict[str, Any]] = []
    first_row = True
    for row in rows:
        name = str(row.get(variant_col, "")).strip()
        if not name:
            continue
        if control_col:
            is_ctrl = _to_bool(row.get(control_col))
        else:
            is_ctrl = first_row
        first_row = False

        visitors = _to_int(row.get(visitors_col), visitors_col)
        variant_obj: dict[str, Any] = {"name": name, "is_control": is_ctrl}
        if is_conversion:
            variant_obj["visitors"] = visitors
            variant_obj["conversions"] = _to_int(row.get(metric_col), metric_col)
        else:
            variant_obj["visitors"] = visitors
            variant_obj["revenue_sum"] = _to_float_or_none(row.get(metric_col))
        variants.append(variant_obj)

    return _build_final_payload(
        variants, primary_metric, defaults, source_type="aggregated", arpu_approximate=arpu_approximate
    )


def _build_final_payload(
    variants: list[dict[str, Any]],
    primary_metric: str,
    defaults: dict[str, Any],
    source_type: str,
    arpu_approximate: bool,
) -> dict[str, Any]:
    return {
        "experiment_name": defaults.get("experiment_name", "table_input_experiment"),
        "method": defaults.get("method", "bayesian"),
        "primary_metric": primary_metric,
        "alpha": defaults.get("alpha", 0.05),
        "sequential_tau": defaults.get("sequential_tau"),
        "decision_thresholds": defaults.get("decision_thresholds"),
        "decision_policy": defaults.get("decision_policy"),
        "guardrails": defaults.get("guardrails", []),
        "samples": defaults.get("samples", 50000),
        "random_seed": defaults.get("random_seed", 7),
        "variants": variants,
        "input_interpretation": {
            "source_type": source_type,
            "row_count": len(variants),
            "arpu_approximate": arpu_approximate,
        },
    }


def build_payload_from_rows(
    rows: list[dict[str, Any]],
    defaults: dict[str, Any] | None = None,
) -> dict[str, Any]:
    shape = detect_input_shape(rows)
    if shape == "row_level":
        return _build_payload_from_row_level(rows, defaults)
    return _build_payload_from_aggregated(rows, defaults)


def build_duration_request_from_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        raise ValueError("Duration input table is empty.")
    row = rows[0]
    return {
        "method": str(row.get("method", "frequentist")).strip().lower(),
        "baseline_rate": float(row["baseline_rate"]),
        "relative_mde": float(row["relative_mde"]),
        "daily_traffic": int(float(row["daily_traffic"])),
        "n_variants": int(float(row.get("n_variants", 2))),
        "alpha": float(row.get("alpha", 0.05)),
        "power": float(row.get("power", 0.8)),
        "max_looks": int(float(row.get("max_looks", 10))),
        "prob_threshold": float(row.get("prob_threshold", 0.95)),
        "max_expected_loss": float(row.get("max_expected_loss", 0.001)),
        "assurance_target": float(row.get("assurance_target", 0.8)),
        "max_days": int(float(row.get("max_days", 60))),
    }
